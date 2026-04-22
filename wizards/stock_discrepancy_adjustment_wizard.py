from odoo import fields, models
from odoo.exceptions import UserError
import base64
import pandas as pd
from io import BytesIO
import logging
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


_logger = logging.getLogger(__name__)

class StockDiscrepancyAdjustmentWizard(models.TransientModel):
    _name = 'stock.discrepancy.adjustment.wizard'
    _description = 'Stock Discrepancy Adjustment'

    file_import = fields.Binary(string="File import")
    name = fields.Char(string="Name")

    def read_excel(self, file):
        # Load the binary file uploaded by the user
        file_content = base64.b64decode(file)
        excel_file = BytesIO(file_content)
        return pd.read_excel(excel_file, sheet_name=0, engine='openpyxl')
        

    def action_import(self):
        warehouse = self.env["stock.warehouse"].search([("name", "=", "Motus")], limit=1)
        if not warehouse:
            raise UserError("Warehouse 'Motus' not found")

        location = warehouse.lot_stock_id
        if not self.file_import:
            raise UserError("Please upload file excel.")
        df = self.read_excel(self.file_import)

        df.columns = [str(c).strip() for c in df.columns]

        # Detect columns
        col_map = {}
        for cname in df.columns:
            low = cname.lower()
            if low == 'sku' or 'product' in low or 'default_code' in low:
                col_map['sku'] = cname
            elif 'lot' in low:
                col_map['lot'] = cname
            elif 'counted' in low:
                col_map['counted'] = cname

        if not all(k in col_map for k in ('sku', 'lot', 'counted')):
            raise UserError("Excel must contain columns: SKU, Lot, Counted.")

        # Prepare models
        Product = self.env['product.product']
        Lot = self.env['stock.lot']
        Quant = self.env['stock.quant']
        VLayer = self.env['stock.valuation.layer']

        # Helpers
        def find_product(sku):
            return Product.search([('default_code', '=', sku)], limit=1)

        def find_lot(product, lot_name):
            return Lot.search([
                ('name', '=', lot_name),
                ('product_id', '=', product.id)
            ], limit=1)

        def get_system_qty(product, lot):
            quants = Quant.search([
                ('product_id', '=', product.id),
                ('lot_id', '=', lot.id),
                ('location_id', '=', location.id),
            ])
            return sum(q.quantity for q in quants)

        def get_unit_cost(product, lot):
            # layers = VLayer.search([
            #     ('product_id', '=', product.id),
            #     ('lot_id', '=', lot.id),
            # ])
            # if layers:
            #     total_val = sum(l.value for l in layers)
            #     total_qty = sum(l.quantity for l in layers)
            #     if total_qty:
            #         return total_val / total_qty
            return product.standard_price 

        def parse_sku(raw_value):
            if not raw_value:
                return ''
            sku = str(raw_value).strip()
            if '-' in sku:
                sku = sku.split('-', 1)[0].strip()
            return sku

        # Parse rows
        results = []
        preview = []

        for idx, row in df.iterrows():
            row_index = int(idx)

            sku = parse_sku(row[col_map['sku']])
            lot_name = str(row[col_map['lot']]).strip() if not pd.isna(row[col_map['lot']]) else ''

            # Skip if counted empty
            if pd.isna(row[col_map['counted']]):
                results.append({
                    'row_index': row_index,
                    'import_status': 'SKIPPED',
                    'error': 'Counted quantity is empty'
                })
                continue

            try:
                counted_qty = float(row[col_map['counted']])
            except Exception as e:
                results.append({
                    'row_index': row_index,
                    'import_status': 'ERROR',
                    'error': f'Invalid Counted Qty: {e}'
                })
                continue

            if not sku:
                results.append({
                    'row_index': row_index,
                    'import_status': 'ERROR',
                    'error': 'Missing SKU'
                })
                continue

            if not lot_name:
                results.append({
                    'row_index': row_index,
                    'import_status': 'ERROR',
                    'error': 'Missing Lot'
                })
                continue

            product = find_product(sku)
            if not product:
                results.append({
                    'row_index': row_index,
                    'import_status': 'ERROR',
                    'error': f'SKU not found: {sku}'
                })
                continue

            lot = find_lot(product, lot_name)
            if not lot:
                results.append({
                    'row_index': row_index,
                    'import_status': 'ERROR',
                    'error': f'Lot not found: {lot_name}'
                })
                continue

            system_qty = get_system_qty(product, lot)
            unit_cost = get_unit_cost(product, lot)
            diff_qty = counted_qty - system_qty

            preview.append({
                'row_index': row_index,
                'sku': sku,
                'product_id': product.id,
                'lot_id': lot.id,
                'lot_name': lot_name,
                'system_qty': system_qty,
                'counted_qty': counted_qty,
                'discrepancy_qty': diff_qty,
                'unit_cost': unit_cost,
                'discrepancy_value': diff_qty * unit_cost,
            })

            status = 'SKIPPED' if abs(diff_qty) < 1e-9 else 'OK'
            results.append({
                'row_index': row_index,
                'import_status': status,
                'error': ''
            })

        # No valid rows
        if not preview:
            raise UserError("No valid rows to import.")

        # Apply inventory via stock.quant
        # ------------------------------------------------------------
        for p in preview:
            quant = Quant.search([
                ('product_id', '=', p['product_id']),
                ('lot_id', '=', p['lot_id']),
                ('location_id', '=', location.id),
            ], limit=1)

            if not quant:
                quant = Quant.create({
                    'product_id': p['product_id'],
                    'lot_id': p['lot_id'],
                    'location_id': location.id,
                })

            if quant.quantity != p['counted_qty']:
                quant.sudo().write({
                    'inventory_quantity': p['counted_qty'],
                    'inventory_quantity_set': True,
                })
                quant.sudo().action_apply_inventory()

        # Build result Excel (single sheet)
        preview_by_idx = {p['row_index']: p for p in preview}
        result_by_idx = {r['row_index']: r for r in results}

        out_rows = []
        for idx, row in df.iterrows():
            row_index = int(idx)
            sku_val = row.get(col_map['sku']) if 'sku' in col_map else ''
            lot_val = row.get(col_map['lot']) if 'lot' in col_map else ''
            counted_val = row.get(col_map['counted']) if 'counted' in col_map else ''

            p = preview_by_idx.get(row_index)
            r = result_by_idx.get(row_index, {})

            out_rows.append({
                'SKU': '' if pd.isna(sku_val) else str(sku_val),
                'Lot_Number': '' if pd.isna(lot_val) else str(lot_val),
                'Counted_Qty': '' if pd.isna(counted_val) else counted_val,
                'System_Qty': '' if not p else p.get('system_qty', ''),
                'Unit_Cost': '' if not p else p.get('unit_cost', ''),
                'Discrepancy_Qty': '' if not p else p.get('discrepancy_qty', ''),
                'Discrepancy_Value': '' if not p else p.get('discrepancy_value', ''),
                'import_status': r.get('import_status', ''),
                'error': r.get('error', ''),
            })

        res_df = pd.DataFrame(out_rows, columns=[
            'SKU',
            'Lot_Number',
            'Counted_Qty',
            'System_Qty',
            'Unit_Cost',
            'Discrepancy_Qty',
            'Discrepancy_Value',
            'import_status',
            'error',
        ])

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            res_df.to_excel(writer, index=False, sheet_name='result')

            ws = writer.sheets.get('result')
            if ws:
                ok_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
                skip_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
                err_fill = PatternFill(fill_type='solid', fgColor='FFC7CE')

                status_col_idx = res_df.columns.get_loc('import_status') + 1
                max_col = ws.max_column
                max_row = ws.max_row

                for r in range(2, max_row + 1):
                    status_val = ws.cell(row=r, column=status_col_idx).value
                    if status_val == 'OK':
                        fill = ok_fill
                    elif status_val == 'SKIPPED':
                        fill = skip_fill
                    elif status_val == 'ERROR':
                        fill = err_fill
                    else:
                        continue

                    for c in range(1, max_col + 1):
                        ws.cell(row=r, column=c).fill = fill

                for col_idx in range(1, max_col + 1):
                    max_len = 0
                    for cell in ws[get_column_letter(col_idx)]:
                        val = cell.value
                        if val is None:
                            continue
                        max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        output.seek(0)
        data = output.read()

        filename = f"stock_counting_{fields.Date.today().strftime('%d-%m-%Y')}_result.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(data),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_get_template(self):
        warehouse = self.env["stock.warehouse"].search([("name", "=", "Motus")], limit=1)
        if not warehouse:
            raise UserError("Warehouse 'Motus' not found")

        location = warehouse.lot_stock_id
        quant_domain = [
            ("location_id", "child_of", location.id),
            ("lot_id", "!=", False),
            ("product_id", "!=", False),
        ]
        quants = self.env["stock.quant"].search(quant_domain)

        rows = []
        seen = set()
        for q in quants:
            key = (q.product_id.id, q.lot_id.id)
            if key in seen:
                continue
            seen.add(key)
            sku_display = ((q.product_id.default_code or "").strip() + " - " + (q.product_id.name or "").strip()).strip()
            rows.append({
                "SKU": sku_display,
                "Lot_Number": q.lot_id.name,
                "Counted": "",
            })

        if not rows:
            rows = [{"SKU": "", "Lot_Number": "", "Counted": ""}]

        df = pd.DataFrame(rows, columns=["SKU", "Lot_Number", "Counted"])

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="template")
        output.seek(0)
        data = output.read()
        attachment = self.env["ir.attachment"].create({
            "name":  f"stock_counting_{fields.Date.today().strftime('%d-%m-%Y')}_template.xlsx",
            "type": "binary",
            "datas": base64.b64encode(data),
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }